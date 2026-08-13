from datetime import datetime
import importlib
import os
import re
import site
import sqlite3
import sys
import time

from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from playwright.sync_api import sync_playwright

# ---------------------------------------------------------------------------
# 1. CONFIGURATION & ENVIRONMENT VARIABLES
# ---------------------------------------------------------------------------
LOGIN_URL = "https://gateway.1-stop.biz/"
SEARCH_URL = (
    "https://gateway.1-stop.biz/VesselSchedule/SearchAdvanced"
    "?searchAction=search&PortOfCall=AUMEL&Terminals=&VesselName="
    "&LineOperator=&LloydsNumber=&SubmitSearchCriteria=Search"
)

# Pull credentials from environment variables (used by GitHub Actions Secrets)
GATEWAY_USERNAME = os.getenv("GATEWAY_USERNAME", "leede")
GATEWAY_PASSWORD = os.getenv("GATEWAY_PASSWORD", "8g*HL#TJkZRYe?4")

# Detect CI/Cloud Environment (GitHub Actions sets CI=true)
IS_CI = os.getenv("CI", "false").lower() == "true"

DB_FILE = "gateway_vessel_schedule.db"
EXCEL_OUTPUT = "Gateway_Vessel_Schedule.xlsx"
PROFILE_DIR = os.path.abspath("./gateway_playwright_profile")

REAL_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0.0.0 Safari/537.36"
)

# ---------------------------------------------------------------------------
# 2. HELPER FUNCTIONS
# ---------------------------------------------------------------------------
def parse_gateway_datetime(val):
    """
    Parses date strings like '04 Oct03:05', '04 Oct 03:05', or '12 Feb16:24'
    into actual Python datetime objects for Excel formatting.
    """
    if not val or pd.isna(val):
        return pd.NaT

    val = str(val).strip()
    if not val or val in ["-", "N/A", "TBA"]:
        return pd.NaT

    clean_val = re.sub(r"([a-zA-Z]{3})\s*(\d{2}:\d{2})", r"\1 \2", val)
    clean_val = re.sub(r"\s+", " ", clean_val)

    curr_year = datetime.now().year

    try:
        return datetime.strptime(f"{clean_val} {curr_year}", "%d %b %H:%M %Y")
    except ValueError:
        pass

    try:
        return datetime.strptime(f"{clean_val} {curr_year}", "%d %b %Y")
    except ValueError:
        pass

    return pd.NaT


def perform_login(page):
    """Executes authentication on the Gateway portal."""
    print(f"Navigating to login page: {LOGIN_URL}...")
    page.goto(LOGIN_URL)
    page.wait_for_timeout(2000)

    if "SignIn" not in page.url and page.locator("#logOnSubmitBtn").count() == 0:
        print("Already authenticated or redirected.")
        return

    print("Filling credentials...")
    user_input = page.locator(
        "#username, input[name='username'], input[type='email'], input[name*='user' i]"
    ).first
    user_input.fill(GATEWAY_USERNAME)

    pass_input = page.locator("#password, input[type='password']").first
    pass_input.fill(GATEWAY_PASSWORD)

    page.wait_for_timeout(500)

    print("Submitting login...")
    if page.locator("#logOnSubmitBtn").count() > 0:
        page.click("#logOnSubmitBtn")
    else:
        page.keyboard.press("Enter")

    print("Waiting for session initialization...")
    page.wait_for_timeout(5000)


def extract_gateway_schedule(page):
    """Navigates to the search URL and parses table #searchResult."""
    print(f"Navigating to Gateway Schedule Search URL...\n{SEARCH_URL}")
    page.goto(SEARCH_URL)
    page.wait_for_timeout(4000)

    if "Just a moment" in page.title() or "challenge" in page.content().lower():
        print("Cloudflare verification detected. Waiting 10 seconds...")
        page.wait_for_timeout(10000)

    if "SignIn" in page.url or page.locator("#logOnSubmitBtn").count() > 0:
        print("Session expired. Re-authenticating...")
        perform_login(page)
        page.goto(SEARCH_URL)
        page.wait_for_timeout(4000)

    print("Locating schedule table '#searchResult'...")
    table_html = None

    for _ in range(15):
        for scope in [page] + page.frames:
            try:
                locator = scope.locator("#searchResult")
                if locator.count() > 0 and locator.first.is_visible():
                    table_html = locator.first.evaluate("el => el.outerHTML")
                    break
            except Exception:
                pass
        if table_html:
            break
        page.wait_for_timeout(1000)

    if not table_html:
        print("Fallback: Extracting full page content for HTML parsing...")
        table_html = page.content()

    soup = BeautifulSoup(table_html, "html.parser")
    target_table = soup.find("table", id="searchResult") or soup.find("table")

    if not target_table:
        print("ERROR: Could not locate table on page.")
        return pd.DataFrame()

    headers = []
    thead = target_table.find("thead")
    if thead:
        for th in thead.find_all("th"):
            text = th.get_text(" ", strip=True)
            if text:
                headers.append(text)

    if not headers:
        headers = [
            "Port or Terminal",
            "Vessel",
            "Lloyds Number",
            "In Voyage",
            "Out Voyage",
            "ETA",
            "ETD",
            "Export Receive Start",
            "Export Haz Receive Start",
            "Export Empty Receive Start",
            "Export Reefer Cutoff",
            "Export Cargo Cutoff",
            "Export Haz Cutoff",
            "Export Empty Cutoff",
            "First Avail.",
            "First Free",
            "Import Storage Start",
            "Berth",
            "Line",
            "Route",
        ]

    rows = []
    tbody = target_table.find("tbody")
    row_elements = (
        tbody.find_all("tr")
        if tbody
        else target_table.find_all("tr", class_=["gridrow", "gridrow_alternate"])
    )

    for tr in row_elements:
        cells = []
        for td in tr.find_all("td"):
            for br in td.find_all("br"):
                br.replace_with(" ")
            cell_text = td.get_text(strip=True)
            cells.append(cell_text)

        if cells:
            if len(cells) > len(headers):
                cells = cells[: len(headers)]
            elif len(cells) < len(headers):
                cells += [""] * (len(headers) - len(cells))
            rows.append(cells)

    df = pd.DataFrame(rows, columns=headers[: len(rows[0])] if rows else None)

    if "Vessel" in df.columns:
        df = df[df["Vessel"].astype(str).str.strip() != ""]

    if "Lloyds Number" in df.columns:
        df["Lloyds Number"] = (
            df["Lloyds Number"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )
    else:
        df["Lloyds Number"] = ""

    if "In Voyage" in df.columns:
        df["In Voyage"] = df["In Voyage"].astype(str).str.strip()
    else:
        df["In Voyage"] = ""

    datetime_cols = [
        "ETA",
        "ETD",
        "Export Receive Start",
        "Export Haz Receive Start",
        "Export Empty Receive Start",
        "Export Reefer Cutoff",
        "Export Cargo Cutoff",
        "Export Haz Cutoff",
        "Export Empty Cutoff",
        "First Avail.",
        "First Free",
        "Import Storage Start",
    ]

    for col in datetime_cols:
        if col in df.columns:
            df[col] = df[col].apply(parse_gateway_datetime)

    df["Scrape Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print(f"Successfully extracted {len(df)} records.")
    return df


def merge_with_historical_data(df_new):
    """
    Merges newly scraped records with existing database records based on
    a composite key of ('Lloyds Number' AND 'In Voyage').
    Preserves historical records while overwriting matching voyages with fresh data.
    """
    if df_new.empty:
        return df_new

    if not os.path.exists(DB_FILE):
        print("No existing database found. Starting fresh historical record.")
        return df_new

    try:
        conn = sqlite3.connect(DB_FILE)
        df_old = pd.read_sql("SELECT * FROM gateway_vessel_schedule", conn)
        conn.close()

        if df_old.empty:
            return df_new

        print(f"Found {len(df_old)} existing historical records in database.")

        # Ensure matching key columns are string types and cleaned
        for frame in [df_new, df_old]:
            if "Lloyds Number" in frame.columns:
                frame["Lloyds Number"] = frame["Lloyds Number"].astype(str).str.strip()
            else:
                frame["Lloyds Number"] = ""

            if "In Voyage" in frame.columns:
                frame["In Voyage"] = frame["In Voyage"].astype(str).str.strip()
            else:
                frame["In Voyage"] = ""

        # Create composite matching key: 'Lloyds Number' + '|' + 'In Voyage'
        df_new["_composite_key"] = df_new["Lloyds Number"] + "|" + df_new["In Voyage"]
        df_old["_composite_key"] = df_old["Lloyds Number"] + "|" + df_old["In Voyage"]

        # Extract set of non-empty composite keys from new scrape batch
        new_keys = set(df_new[df_new["_composite_key"] != "|"]["_composite_key"])

        # Retain old records whose composite keys are NOT in the new batch
        df_old_retained = df_old[~df_old["_composite_key"].isin(new_keys)].copy()

        # Combine old retained records with all new records
        df_combined = pd.concat([df_old_retained, df_new], ignore_index=True)

        # Drop temporary composite key
        df_combined.drop(columns=["_composite_key"], inplace=True, errors="ignore")
        if "_composite_key" in df_new.columns:
            df_new.drop(columns=["_composite_key"], inplace=True)

        datetime_cols = [
            "ETA",
            "ETD",
            "Export Receive Start",
            "Export Haz Receive Start",
            "Export Empty Receive Start",
            "Export Reefer Cutoff",
            "Export Cargo Cutoff",
            "Export Haz Cutoff",
            "Export Empty Cutoff",
            "First Avail.",
            "First Free",
            "Import Storage Start",
        ]
        for col in datetime_cols:
            if col in df_combined.columns:
                df_combined[col] = pd.to_datetime(df_combined[col], errors="coerce")

        print(
            f"Historical Merge Complete (Key: Lloyds Number + In Voyage): "
            f"Preserved {len(df_old_retained)} old records, updated/added {len(df_new)} records (Total: {len(df_combined)} rows)."
        )
        return df_combined

    except Exception as e:
        print(f"Warning: Could not read existing database for merge ({e}). Proceeding with current scrape only.")
        return df_new


def export_to_excel_and_sqlite(df_new):
    """Merges historical data, then saves to SQLite and exports formatted Excel Table."""
    if df_new.empty:
        print("No data extracted. Skipping export.")
        return

    df_final = merge_with_historical_data(df_new)

    print(f"\n1. Saving {len(df_final)} total records to SQLite database: {DB_FILE}...")
    conn = sqlite3.connect(DB_FILE)
    df_to_sql = df_final.copy()

    for col in df_to_sql.select_dtypes(include=["datetime64[ns]", "datetime64"]):
        df_to_sql[col] = df_to_sql[col].dt.strftime("%Y-%m-%d %H:%M:%S")

    df_to_sql.to_sql("gateway_vessel_schedule", conn, if_exists="replace", index=False)
    conn.close()

    print(f"2. Exporting formatted Excel Table to: {EXCEL_OUTPUT}...")

    with pd.ExcelWriter(
        EXCEL_OUTPUT,
        engine="openpyxl",
        datetime_format="YYYY-MM-DD HH:MM",
        date_format="YYYY-MM-DD",
    ) as writer:
        sheet_name = "Gateway Schedule"
        df_final.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]
        max_row = len(df_final) + 1
        max_col = len(df_final.columns)
        col_letter = get_column_letter(max_col)

        tab_range = f"A1:{col_letter}{max_row}"
        table = Table(displayName="GatewayVesselScheduleList", ref=tab_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        ws.add_table(table)

        for col in ws.columns:
            max_len = 0
            for cell in col:
                if isinstance(cell.value, datetime):
                    val_str = cell.value.strftime("%Y-%m-%d %H:%M")
                else:
                    val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))

            col_letter_idx = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter_idx].width = max(max_len + 3, 12)

    print(f"\nSUCCESS: Exported {len(df_final)} historical and current rows to Excel and SQLite!")


def run():
    print(f"Starting Playwright Context (Headless Mode = {IS_CI})...")
    with sync_playwright() as p:
        stealth_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--no-first-run",
        ]

        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=PROFILE_DIR,
                headless=IS_CI,
                user_agent=REAL_USER_AGENT,
                viewport={"width": 1400, "height": 900},
                ignore_default_args=["--enable-automation"],
                args=stealth_args,
            )
        except Exception:
            context = p.chromium.launch_persistent_context(
                user_data_dir=PROFILE_DIR,
                headless=IS_CI,
                user_agent=REAL_USER_AGENT,
                viewport={"width": 1400, "height": 900},
                ignore_default_args=["--enable-automation"],
                args=stealth_args,
            )

        page = context.pages[0] if context.pages else context.new_page()

        page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
        )

        perform_login(page)
        df_schedule = extract_gateway_schedule(page)
        export_to_excel_and_sqlite(df_schedule)

        context.close()


if __name__ == "__main__":
    try:
        run()
    except Exception as e:
        print(f"\nAN ERROR OCCURRED: {e}")
