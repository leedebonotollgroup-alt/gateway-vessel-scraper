from datetime import datetime
import importlib
import os
import re
import site
import sqlite3
import subprocess
import sys
import time

# ---------------------------------------------------------------------------
# 1. DEPENDENCY CHECK & AUTO-INSTALL
# ---------------------------------------------------------------------------
user_site = site.getusersitepackages()
if user_site not in sys.path:
    sys.path.insert(0, user_site)

REQUIRED_PACKAGES = {
    "pandas": "pandas",
    "openpyxl": "openpyxl",
    "playwright": "playwright",
    "bs4": "beautifulsoup4",
}


def install_requirements():
    python_exe = sys.executable.replace("pythonw.exe", "python.exe")
    for module_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(module_name)
        except ImportError:
            print(f"Installing missing package: {pip_name}...")
            subprocess.check_call(
                [
                    python_exe,
                    "-m",
                    "pip",
                    "install",
                    "--user",
                    "--trusted-host",
                    "pypi.org",
                    "--trusted-host",
                    "files.pythonhosted.org",
                    pip_name,
                ]
            )
    importlib.invalidate_caches()


install_requirements()

from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from playwright.sync_api import sync_playwright

# ---------------------------------------------------------------------------
# 2. CONFIGURATION & ENVIRONMENT VARIABLES
# ---------------------------------------------------------------------------
LOGIN_URL = "https://gateway.1-stop.biz/"
SEARCH_URL = (
    "https://gateway.1-stop.biz/VesselSchedule/SearchAdvanced"
    "?searchAction=search&PortOfCall=AUMEL&Terminals=&VesselName="
    "&LineOperator=&LloydsNumber=&SubmitSearchCriteria=Search"
)

# Pull credentials from environment variables (used by GitHub Actions Secrets), fallback to defaults
USERNAME = os.getenv("GATEWAY_USERNAME", "leede")
PASSWORD = os.getenv("GATEWAY_PASSWORD", "8g*HL#TJkZRYe?4")

# Detect CI/Cloud Environment (GitHub Actions sets CI=true)
IS_CI = os.getenv("CI", "false").lower() == "true"

DB_FILE = "gateway_vessel_schedule.db"
EXCEL_OUTPUT = "Gateway_Vessel_Schedule.xlsx"
PROFILE_DIR = os.path.abspath("./gateway_playwright_profile")

REAL_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0.0.0 Safari/537.36"
)

# ---------------------------------------------------------------------------
# 3. HELPER FUNCTIONS
# ---------------------------------------------------------------------------
def parse_gateway_datetime(val):
    """
    Parses date strings like '04 Oct03:05', '04 Oct 03:05', or '12 Feb16:24'
    into actual Python datetime objects for Excel formatting.
    """
    if not val or pd.isna(val):
        return pd.NaT

    val = str(val).strip()
    if not val:
        return pd.NaT

    # Standardize format by inserting a space between month abbreviation and time
    clean_val = re.sub(r"([a-zA-Z]{3})\s*(\d{2}:\d{2})", r"\1 \2", val)
    clean_val = re.sub(r"\s+", " ", clean_val)

    curr_year = datetime.now().year

    # Format 1: '04 Oct 03:05'
    try:
        return datetime.strptime(f"{clean_val} {curr_year}", "%d %b %H:%M %Y")
    except ValueError:
        pass

    # Format 2: '04 Oct' (Date only)
    try:
        return datetime.strptime(f"{clean_val} {curr_year}", "%d %b %Y")
    except ValueError:
        pass

    return pd.NaT


def perform_login(page):
    """Executes authentication on the Gateway portal."""
    print(f"Navigating to login page: {LOGIN_URL}...")
    page.goto(LOGIN_URL)
    page.wait_for_timeout(2000)

    if "SignIn" not in page.url and page.locator("#logOnSubmitBtn").count() == 0:
        print("Already authenticated or redirected.")
        return

    print("Filling credentials...")
    user_input = page.locator(
        "#username, input[name='username'], input[type='email'], input[name*='user' i]"
    ).first
    user_input.fill(USERNAME)

    pass_input = page.locator("#password, input[type='password']").first
    pass_input.fill(PASSWORD)

    page.wait_for_timeout(500)

    print("Submitting login...")
    if page.locator("#logOnSubmitBtn").count() > 0:
        page.click("#logOnSubmitBtn")
    else:
        page.keyboard.press("Enter")

    print("Waiting for session initialization...")
    page.wait_for_timeout(5000)


def extract_gateway_schedule(page):
    """Navigates to the search URL and parses table #searchResult."""
    print(f"Navigating to Gateway Schedule Search URL...\n{SEARCH_URL}")
    page.goto(SEARCH_URL)
    page.wait_for_timeout(4000)

    if "Just a moment" in page.title() or "challenge" in page.content().lower():
        print("Cloudflare verification detected. Waiting 10 seconds...")
        page.wait_for_timeout(10000)

    if "SignIn" in page.url or page.locator("#logOnSubmitBtn").count() > 0:
        print("Session expired. Re-authenticating...")
        perform_login(page)
        page.goto(SEARCH_URL)
        page.wait_for_timeout(4000)

    print("Locating table '#searchResult'...")
    table_html = None

    for _ in range(15):
        for scope in [page] + page.frames:
            try:
                locator = scope.locator("#searchResult")
                if locator.count() > 0 and locator.first.is_visible():
                    table_html = locator.first.evaluate("el => el.outerHTML")
                    break
            except Exception:
                pass
        if table_html:
            break
        page.wait_for_timeout(1000)

    if not table_html:
        print("Fallback: Extracting full page content for HTML parsing...")
        table_html = page.content()

    soup = BeautifulSoup(table_html, "html.parser")
    target_table = soup.find("table", id="searchResult") or soup.find("table")

    if not target_table:
        print("ERROR: Could not locate table on page.")
        return pd.DataFrame()

    headers = []
    thead = target_table.find("thead")
    if thead:
        for th in thead.find_all("th"):
            text = th.get_text(" ", strip=True)
            if text:
                headers.append(text)

    if not headers:
        headers = [
            "Port or Terminal",
            "Vessel",
            "Lloyds Number",
            "In Voyage",
            "Out Voyage",
            "ETA",
            "ETD",
            "Export Receive Start",
            "Export Haz Receive Start",
            "Export Empty Receive Start",
            "Export Reefer Cutoff",
            "Export Cargo Cutoff",
            "Export Haz Cutoff",
            "Export Empty Cutoff",
            "First Avail.",
            "First Free",
            "Import Storage Start",
            "Berth",
            "Line",
            "Route",
        ]

    rows = []
    tbody = target_table.find("tbody")
    row_elements = (
        tbody.find_all("tr")
        if tbody
        else target_table.find_all("tr", class_=["gridrow", "gridrow_alternate"])
    )

    for tr in row_elements:
        cells = []
        for td in tr.find_all("td"):
            for br in td.find_all("br"):
                br.replace_with(" ")
            cell_text = td.get_text(strip=True)
            cells.append(cell_text)

        if cells:
            if len(cells) > len(headers):
                cells = cells[: len(headers)]
            elif len(cells) < len(headers):
                cells += [""] * (len(headers) - len(cells))
            rows.append(cells)

    df = pd.DataFrame(rows, columns=headers[: len(rows[0])] if rows else None)

    if "Vessel" in df.columns:
        df = df[df["Vessel"].astype(str).str.strip() != ""]

    # Convert Date/Time columns to actual Python datetime objects
    datetime_cols = [
        "ETA",
        "ETD",
        "Export Receive Start",
        "Export Haz Receive Start",
        "Export Empty Receive Start",
        "Export Reefer Cutoff",
        "Export Cargo Cutoff",
        "Export Haz Cutoff",
        "Export Empty Cutoff",
        "First Avail.",
        "First Free",
        "Import Storage Start",
    ]

    for col in datetime_cols:
        if col in df.columns:
            df[col] = df[col].apply(parse_gateway_datetime)

    return df


def export_to_excel_and_sqlite(df):
    """Exports DataFrame to SQLite database and formats output as an Excel Table."""
    if df.empty:
        print("No data extracted. Skipping export.")
        return

    print(f"\n1. Saving {len(df)} records to SQLite database: {DB_FILE}...")
    conn = sqlite3.connect(DB_FILE)
    df.to_sql("gateway_vessel_schedule", conn, if_exists="replace", index=False)
    conn.close()

    print(f"2. Exporting formatted Excel Table to: {EXCEL_OUTPUT}...")

    with pd.ExcelWriter(
        EXCEL_OUTPUT,
        engine="openpyxl",
        datetime_format="YYYY-MM-DD HH:MM",
        date_format="YYYY-MM-DD",
    ) as writer:
        sheet_name = "Gateway Schedule"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]
        max_row = len(df) + 1
        max_col = len(df.columns)
        col_letter = get_column_letter(max_col)

        # Apply Native Excel Table Formatting
        tab_range = f"A1:{col_letter}{max_row}"
        table = Table(displayName="GatewayVesselScheduleList", ref=tab_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        ws.add_table(table)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if isinstance(cell.value, datetime):
                    val_str = cell.value.strftime("%Y-%m-%d %H:%M")
                else:
                    val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))

            col_letter_idx = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter_idx].width = max(max_len + 3, 12)

    print(f"\nSUCCESS: Scraped {len(df)} rows with formatted DateTime fields!")


# ---------------------------------------------------------------------------
# 4. MAIN EXECUTION ROUTINE
# ---------------------------------------------------------------------------
def run():
    print(f"Starting Playwright Context (Headless Mode = {IS_CI})...")
    with sync_playwright() as p:
        stealth_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--no-first-run",
        ]

        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=PROFILE_DIR,
                headless=IS_CI,  # True on GitHub Actions, False locally
                user_agent=REAL_USER_AGENT,
                viewport={"width": 1400, "height": 900},
                ignore_default_args=["--enable-automation"],
                args=stealth_args,
            )
        except Exception:
            context = p.chromium.launch_persistent_context(
                user_data_dir=PROFILE_DIR,
                headless=IS_CI,
                user_agent=REAL_USER_AGENT,
                viewport={"width": 1400, "height": 900},
                ignore_default_args=["--enable-automation"],
                args=stealth_args,
            )

        page = context.pages[0] if context.pages else context.new_page()

        page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
        )

        perform_login(page)
        df_schedule = extract_gateway_schedule(page)
        export_to_excel_and_sqlite(df_schedule)

        context.close()


if __name__ == "__main__":
    try:
        run()
    except Exception as e:
        print(f"\nAN ERROR OCCURRED: {e}")
